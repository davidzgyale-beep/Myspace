import pandas as pd
from openpyxl import load_workbook

from edgar_exporter.excel_writer import write_excel


def test_write_excel_creates_all_required_sheets(tmp_path):
    statements = {
        "Income Statement": pd.DataFrame({"Line Item": ["Revenue", "EPS Basic"], "FY2022": [1000, 1.23]}),
        "Balance Sheet": pd.DataFrame({"Line Item": ["Total Assets"], "FY2022": [5000]}),
        "Cash Flow Statement": pd.DataFrame({"Line Item": ["Net Change in Cash"], "FY2022": [100]}),
    }
    raw_facts_df = pd.DataFrame(
        {
            "taxonomy": ["us-gaap"],
            "tag": ["Revenues"],
            "label": ["Revenues"],
            "description": ["desc"],
            "unit": ["USD"],
            "value": [1000],
            "fiscal_year": [2022],
            "fiscal_period": ["FY"],
            "form": ["10-K"],
            "filed": ["2023-01-01"],
            "frame": [None],
            "accession_number": ["ACC-1"],
            "start": ["2022-01-01"],
            "end": ["2022-12-31"],
        }
    )
    quality_df = pd.DataFrame(
        {
            "Statement": ["Income Statement"],
            "Statement Line Item": ["Revenue"],
            "Period": ["FY2022"],
            "Status": ["OK"],
            "Source Tag": ["Revenues"],
            "Fallback Tag Used": [None],
            "Unit": ["USD"],
            "Filed Date": ["2023-01-01"],
            "Accession Number": ["ACC-1"],
        }
    )
    quality_summary_df = pd.DataFrame({"Metric": ["Total"], "Value": [1]})
    metadata = {
        "Company Name": "Apple Inc.",
        "Ticker/Input": "AAPL",
        "CIK": "0000320193",
        "Period Type": "annual",
        "Start Year": 2020,
        "End Year": 2022,
        "Generated At (UTC)": "2024-01-01 00:00:00",
        "Data Source": "SEC EDGAR",
    }

    output_path = tmp_path / "test_output.xlsx"
    write_excel(
        output_path=output_path,
        statements=statements,
        raw_facts_df=raw_facts_df,
        quality_df=quality_df,
        quality_summary_df=quality_summary_df,
        metadata=metadata,
    )

    assert output_path.exists()

    wb = load_workbook(output_path)
    expected_sheets = {
        "Income Statement",
        "Balance Sheet",
        "Cash Flow Statement",
        "Raw Facts",
        "Metadata",
        "Data Quality Check",
    }
    assert expected_sheets.issubset(set(wb.sheetnames))

    ws = wb["Income Statement"]
    assert ws["A1"].value == "Line Item"
    assert ws["B1"].value == "FY2022"
    assert ws.freeze_panes == "B2"
    # EPS row (row 3) should use a decimal format, Revenue row (row 2) a whole-number one.
    assert "0.00" in ws["B3"].number_format
    assert "0.00" not in ws["B2"].number_format

    ws_meta = wb["Metadata"]
    assert ws_meta["A1"].value == "Field"

    ws_quality = wb["Data Quality Check"]
    assert ws_quality["A1"].value == "Statement"

    # Legend is written next to each statement table.
    assert ws.cell(row=1, column=4).value == "Legend"


def test_write_excel_shades_fallback_and_missing_cells(tmp_path):
    statements = {
        "Income Statement": pd.DataFrame(
            {"Line Item": ["Revenue", "Interest Expense"], "FY2022": [1000, None]}
        ),
    }
    raw_facts_df = pd.DataFrame(columns=["taxonomy", "tag", "value"])
    quality_by_statement = {
        "Income Statement": pd.DataFrame(
            {
                "Statement Line Item": ["Revenue", "Interest Expense"],
                "Period": ["FY2022", "FY2022"],
                "Status": ["Fallback Used", "Missing"],
                "Source Tag": ["SalesRevenueNet", None],
                "Fallback Tag Used": ["SalesRevenueNet", None],
                "Unit": ["USD", None],
                "Filed Date": ["2023-01-01", None],
                "Accession Number": ["ACC-1", None],
            }
        )
    }
    quality_df = quality_by_statement["Income Statement"].copy()
    quality_df.insert(0, "Statement", "Income Statement")
    quality_summary_df = pd.DataFrame({"Metric": ["Total"], "Value": [2]})
    metadata = {"Company Name": "Test Co."}

    output_path = tmp_path / "shading.xlsx"
    write_excel(
        output_path=output_path,
        statements=statements,
        raw_facts_df=raw_facts_df,
        quality_df=quality_df,
        quality_summary_df=quality_summary_df,
        metadata=metadata,
        quality_by_statement=quality_by_statement,
    )

    wb = load_workbook(output_path)
    ws = wb["Income Statement"]
    assert ws["B2"].fill.start_color.rgb == "FFFFF59D"  # Revenue: fallback -> yellow
    assert ws["B3"].fill.start_color.rgb == "FFFFCDD2"  # Interest Expense: missing -> red


def test_write_excel_includes_reconciliation_and_anomaly_sections(tmp_path):
    statements = {
        "Income Statement": pd.DataFrame(
            {"Line Item": ["Net Income"], "FY2021": [100], "FY2022": [1000]}
        ),
        "Balance Sheet": pd.DataFrame(
            {
                "Line Item": ["Total Assets", "Total Liabilities and Equity", "Cash and Cash Equivalents"],
                "FY2021": [500, 500, 100],
                "FY2022": [600, 700, 200],  # deliberately unbalanced in FY2022
            }
        ),
        "Cash Flow Statement": pd.DataFrame(
            {
                "Line Item": ["Net Income", "Net Change in Cash", "Cash at End of Period"],
                "FY2021": [100, 10, 100],
                "FY2022": [100, 100, 200],
            }
        ),
    }
    raw_facts_df = pd.DataFrame(columns=["taxonomy", "tag", "value"])
    quality_df = pd.DataFrame(columns=[
        "Statement", "Statement Line Item", "Period", "Status", "Source Tag",
        "Fallback Tag Used", "Unit", "Filed Date", "Accession Number",
    ])
    quality_summary_df = pd.DataFrame({"Metric": ["Total"], "Value": [0]})
    metadata = {"Company Name": "Test Co."}

    from edgar_exporter.data_quality import build_reconciliation_checks, detect_anomalies

    reconciliation_df = build_reconciliation_checks(statements)
    anomalies_df = detect_anomalies(statements)

    output_path = tmp_path / "recon.xlsx"
    write_excel(
        output_path=output_path,
        statements=statements,
        raw_facts_df=raw_facts_df,
        quality_df=quality_df,
        quality_summary_df=quality_summary_df,
        metadata=metadata,
        reconciliation_df=reconciliation_df,
        anomalies_df=anomalies_df,
    )

    wb = load_workbook(output_path)
    ws = wb["Data Quality Check"]
    col_a_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "Reconciliation Checks" in col_a_values
    assert "Anomaly Flags (large period-over-period changes)" in col_a_values
    assert "Total Assets = Total Liabilities and Equity" in col_a_values
