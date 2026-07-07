"""Render statements + raw facts + metadata + data quality into a formatted .xlsx workbook."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict, Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

NUMBER_FORMAT = "#,##0;(#,##0)"
DECIMAL_FORMAT = "#,##0.00;(#,##0.00)"
PERCENT_FORMAT = "0.0%"

# Line items reported as a ratio (XBRL unit "pure") rather than a dollar or
# share amount, and so need percentage formatting instead of NUMBER_FORMAT.
PERCENT_LINE_ITEMS = {"Effective Tax Rate"}

FALLBACK_FILL = PatternFill(start_color="FFFFF59D", end_color="FFFFF59D", fill_type="solid")
MISSING_FILL = PatternFill(start_color="FFFFCDD2", end_color="FFFFCDD2", fill_type="solid")
MISMATCH_FILL = PatternFill(start_color="FFFFCDD2", end_color="FFFFCDD2", fill_type="solid")
ANOMALY_FILL = PatternFill(start_color="FFFFF59D", end_color="FFFFF59D", fill_type="solid")


def _autofit_columns(worksheet, df: pd.DataFrame, min_width: int = 10, max_width: int = 60) -> None:
    """Widen columns to fit content. Never shrinks a column another section already widened."""
    for idx, col in enumerate(df.columns, start=1):
        series_len = df[col].map(lambda v: len(str(v))).max() if len(df) else 0
        header_len = len(str(col))
        width = min(max(series_len, header_len, min_width) + 2, max_width)
        col_letter = get_column_letter(idx)
        current = worksheet.column_dimensions[col_letter].width
        if current is None or width > current:
            worksheet.column_dimensions[col_letter].width = width


def _style_header(worksheet, ncols: int, header_row: int = 1) -> None:
    for col_idx in range(1, ncols + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


def _write_section_title(worksheet, row: int, title: str) -> None:
    cell = worksheet.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, italic=True)


def _apply_statement_number_formats(
    worksheet, df: pd.DataFrame, quality_lookup: Optional[Dict[tuple, str]] = None
) -> None:
    """Set number formats per line item, and (if a quality lookup is given) shade cells
    whose value came from a fallback tag (yellow) or is missing (red) so data quality
    is visible directly on the statement, not just on the Data Quality Check sheet.
    """
    period_cols = [c for c in df.columns if c != "Line Item"]
    # EPS rows use decimals; ratio rows (e.g. Effective Tax Rate) use percentages;
    # everything else (dollar amounts, share counts) stays a whole number.
    line_items = df["Line Item"].astype(str)
    decimal_rows = set(df.index[line_items.str.startswith("EPS")].tolist())
    percent_rows = set(df.index[line_items.isin(PERCENT_LINE_ITEMS)].tolist())

    for col_name in period_cols:
        col_idx = df.columns.get_loc(col_name) + 1
        for row_idx0 in range(len(df)):
            cell = worksheet.cell(row=row_idx0 + 2, column=col_idx)
            if row_idx0 in percent_rows:
                cell.number_format = PERCENT_FORMAT
            elif row_idx0 in decimal_rows:
                cell.number_format = DECIMAL_FORMAT
            else:
                cell.number_format = NUMBER_FORMAT

            if quality_lookup is not None:
                status = quality_lookup.get((line_items.iloc[row_idx0], col_name))
                if status == "Fallback Used":
                    cell.fill = FALLBACK_FILL
                elif status == "Missing":
                    cell.fill = MISSING_FILL


def _write_legend(worksheet, df: pd.DataFrame) -> None:
    legend_col = len(df.columns) + 2
    worksheet.cell(row=1, column=legend_col, value="Legend").font = Font(bold=True)
    worksheet.cell(row=2, column=legend_col).fill = FALLBACK_FILL
    worksheet.cell(row=2, column=legend_col + 1, value="Fallback tag used")
    worksheet.cell(row=3, column=legend_col).fill = MISSING_FILL
    worksheet.cell(row=3, column=legend_col + 1, value="Missing (not reported by filer)")
    worksheet.column_dimensions[get_column_letter(legend_col + 1)].width = 28


def _apply_number_format_to_column(worksheet, df: pd.DataFrame, column: str, fmt: str = NUMBER_FORMAT) -> None:
    if column not in df.columns:
        return
    col_idx = df.columns.get_loc(column) + 1
    for row_idx in range(2, len(df) + 2):
        worksheet.cell(row=row_idx, column=col_idx).number_format = fmt


def _shade_status_column(worksheet, df: pd.DataFrame, header_row: int, status_column: str) -> None:
    """Shade rows in a Status column ("Mismatch" -> red) so problems jump out visually."""
    if status_column not in df.columns:
        return
    col_idx = df.columns.get_loc(status_column) + 1
    for row_idx0 in range(len(df)):
        value = df.iloc[row_idx0][status_column]
        if value == "Mismatch":
            worksheet.cell(row=header_row + 1 + row_idx0, column=col_idx).fill = MISMATCH_FILL


def write_excel(
    output_path,
    statements: Dict[str, pd.DataFrame],
    raw_facts_df: pd.DataFrame,
    quality_df: pd.DataFrame,
    quality_summary_df: pd.DataFrame,
    metadata: dict,
    quality_by_statement: Optional[Dict[str, pd.DataFrame]] = None,
    reconciliation_df: Optional[pd.DataFrame] = None,
    anomalies_df: Optional[pd.DataFrame] = None,
) -> None:
    """Write the full workbook: one sheet per statement, plus Raw Facts, Metadata,
    and Data Quality Check (per-cell provenance, summary, cross-statement
    reconciliation checks, and period-over-period anomaly flags).
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    quality_by_statement = quality_by_statement or {}
    reconciliation_df = reconciliation_df if reconciliation_df is not None else pd.DataFrame()
    anomalies_df = anomalies_df if anomalies_df is not None else pd.DataFrame()

    metadata_df = pd.DataFrame(list(metadata.items()), columns=["Field", "Value"])

    # Per-statement (line item, period) -> Status lookup, used to shade statement cells.
    quality_lookups: Dict[str, Dict[tuple, str]] = {}
    for statement_name, qdf in quality_by_statement.items():
        if qdf.empty:
            quality_lookups[statement_name] = {}
            continue
        quality_lookups[statement_name] = {
            (r["Statement Line Item"], r["Period"]): r["Status"] for _, r in qdf.iterrows()
        }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in statements.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

        raw_facts_df.to_excel(writer, sheet_name="Raw Facts", index=False)
        metadata_df.to_excel(writer, sheet_name="Metadata", index=False)

        # --- Data Quality Check sheet: four stacked sections ---
        quality_df.to_excel(writer, sheet_name="Data Quality Check", index=False, startrow=0)

        summary_start = len(quality_df) + 3
        quality_summary_df.to_excel(
            writer, sheet_name="Data Quality Check", index=False, startrow=summary_start
        )
        summary_header_row = summary_start + 1
        summary_last_row = summary_header_row + len(quality_summary_df)

        recon_title_row = summary_last_row + 3
        recon_header_row = recon_title_row + 1
        reconciliation_df.to_excel(
            writer, sheet_name="Data Quality Check", index=False, startrow=recon_header_row - 1
        )
        recon_last_row = recon_header_row + len(reconciliation_df)

        anomaly_title_row = recon_last_row + 3
        anomaly_header_row = anomaly_title_row + 1
        anomalies_df.to_excel(
            writer, sheet_name="Data Quality Check", index=False, startrow=anomaly_header_row - 1
        )

        # --- Formatting pass ---
        for sheet_name, df in statements.items():
            ws = writer.sheets[sheet_name[:31]]
            ws.freeze_panes = "B2"
            _style_header(ws, len(df.columns))
            _autofit_columns(ws, df)
            _apply_statement_number_formats(ws, df, quality_lookups.get(sheet_name))
            _write_legend(ws, df)

        ws_raw = writer.sheets["Raw Facts"]
        ws_raw.freeze_panes = "A2"
        _style_header(ws_raw, len(raw_facts_df.columns))
        _autofit_columns(ws_raw, raw_facts_df)
        _apply_number_format_to_column(ws_raw, raw_facts_df, "value")

        ws_meta = writer.sheets["Metadata"]
        _style_header(ws_meta, len(metadata_df.columns))
        _autofit_columns(ws_meta, metadata_df)

        ws_quality = writer.sheets["Data Quality Check"]
        ws_quality.freeze_panes = "A2"
        _style_header(ws_quality, len(quality_df.columns))
        _autofit_columns(ws_quality, quality_df)

        if not quality_summary_df.empty:
            _style_header(ws_quality, len(quality_summary_df.columns), header_row=summary_header_row)
            _autofit_columns(ws_quality, quality_summary_df)

        _write_section_title(ws_quality, recon_title_row, "Reconciliation Checks")
        if not reconciliation_df.empty:
            _style_header(ws_quality, len(reconciliation_df.columns), header_row=recon_header_row)
            _autofit_columns(ws_quality, reconciliation_df)
            _shade_status_column(ws_quality, reconciliation_df, recon_header_row, "Status")

        _write_section_title(ws_quality, anomaly_title_row, "Anomaly Flags (large period-over-period changes)")
        if not anomalies_df.empty:
            _style_header(ws_quality, len(anomalies_df.columns), header_row=anomaly_header_row)
            _autofit_columns(ws_quality, anomalies_df)
            for row_idx0 in range(len(anomalies_df)):
                ws_quality.cell(row=anomaly_header_row + 1 + row_idx0, column=1).fill = ANOMALY_FILL

    logger.info("Workbook written to %s", output_path)
